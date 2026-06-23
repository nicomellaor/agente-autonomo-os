import os
import csv
from datetime import datetime
from fpdf import FPDF
from docx import Document
import openpyxl
import re


FORMATOS_PERMITIDOS = {"txt", "md", "pdf", "docx", "xlsx", "csv"}

DIRECTORIO_SALIDA = os.path.expanduser("~/Documents/agente-os")


def extraer_formato(contexto: str) -> str:
    for fmt in FORMATOS_PERMITIDOS:
        if fmt in contexto.lower():
            return fmt
    return "txt"


def extraer_nombre(contexto: str) -> str:
    for token in contexto.split():
        for fmt in FORMATOS_PERMITIDOS:
            if token.endswith(f".{fmt}"):
                return token
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"archivo_{timestamp}"


def generar_contenido(contexto: str) -> str:
    patrones = [
        r"que diga[:\s]+(.+)",
        r"con el texto[:\s]+(.+)",
        r"que contenga[:\s]+(.+)",
        r"escribe[:\s]+(.+)",
        r"con el mensaje[:\s]+(.+)",
        r"con el contenido[:\s]+(.+)"
    ]
    for patron in patrones:
        match = re.search(patron, contexto, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    # Fallback: devuelve el contexto limpio sin la parte de la instrucción
    contexto_limpio = re.sub(
        r"(genera|crea|haz?)\s+(un\s+)?(archivo\s+)?\S+\s*",
        "",
        contexto,
        flags=re.IGNORECASE
    ).strip()
    return contexto_limpio or contexto


def generar_archivo(contexto: str, contenido: str = "") -> None:
    os.makedirs(DIRECTORIO_SALIDA, exist_ok=True)

    fmt = extraer_formato(contexto)
    nombre = extraer_nombre(contexto)
    if not nombre.endswith(f".{fmt}"):
        nombre = f"{nombre}.{fmt}"

    ruta = os.path.join(DIRECTORIO_SALIDA, nombre)

    if fmt in ("txt", "md"):
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(contenido)

    elif fmt == "csv":
        filas = [linea.split(",") for linea in contenido.splitlines()]
        with open(ruta, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(filas)

    elif fmt == "docx":
        doc = Document()
        for linea in contenido.splitlines():
            doc.add_paragraph(linea)
        doc.save(ruta)

    elif fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, linea in enumerate(contenido.splitlines(), start=1):
            for j, celda in enumerate(linea.split(","), start=1):
                ws.cell(row=i, column=j, value=celda.strip())
        wb.save(ruta)

    elif fmt == "pdf":
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        for linea in contenido.splitlines():
            pdf.cell(0, 10, linea, ln=True)
        pdf.output(ruta)

    print(f"\nArchivo generado: {ruta}")